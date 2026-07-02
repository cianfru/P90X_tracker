#!/usr/bin/env python3
"""
P90X.xlsx -> normalized JSON for the offline logger.

Emits: { "exercises": [...], "templates": [...], "sessions": [...] }
matching the app's data model. Encapsulates the spreadsheet quirks documented
in the build brief (swapped dates, reps x kg, a+b, typed modifiers, rounds).

Usage: python import_p90x.py P90X.xlsx > seed.json
"""
import sys, re, json, datetime, uuid, collections
import openpyxl

# Exercises that are pulls (modifier semantics: L/X = harder, band = lighter travel).
PULL = {"Chin ups", "Wide pull", "Pull ups", "Pullups", "Closed grip", "Close grip",
        "Close grip pull", "Switch grip", "Swith grip", "Zip kip pull",
        "Zip kip chin", "Admins chin ups"}

# Canonicalize obvious aliases / typos so data never fragments.
# The last block (below "Heavy P.") was confirmed with the owner in Phase 2:
# pure spelling/punctuation variants of the same movement.
ALIASES = {
    "Closed grip": "Close grip", "Close grip pull": "Close grip",
    "Pullups": "Pull ups", "Swith grip": "Switch grip",
    "Heavy p": "Heavy P.", "Heavy P": "Heavy P.",
    "Declined Push": "Decline push",
    "Side tris rise": "Side tri rise",
    "Balanced curls": "Balance curls",
    "Laying down tris": "Lay down tris",
    "Lawn": "Lawn.",
    # NB: 'Renegade roman' (Renegade Roman Curl) and 'Renegade row' (Med Ball
    # Renegade Row) are DISTINCT moves on the official V-Sculpt sheet — not merged.
}
def canon(name): return ALIASES.get(name, name)

# Exercises logged under their own row that are really a standard move done with
# a harder modifier — fold into the base move and force the modifier so the
# baseline + harder-variant analytics stay correct (Phase 2 decision).
FORCE_MOVE_MOD = {
    "L pullups": ("Pull ups", "L_sit"),
}

def parse_date(v):
    """Row-1 dates: dd/mm/yy strings, OR Excel datetimes with day/month SWAPPED."""
    if isinstance(v, datetime.datetime):
        try:
            return datetime.date(v.year, v.day, v.month)   # swap back
        except ValueError:
            return v.date()
    if isinstance(v, str):
        m = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", v.strip())
        if m:
            d, mo, y = int(m[1]), int(m[2]), int(m[3])
            y = y + 2000 if y < 100 else y
            try:
                return datetime.date(y, mo, d)
            except ValueError:
                return None
    return None

def rxw(s): return re.search(r"(\d+)\s*x\s*(\d+)", s.lower())

# --- Per-day session metadata (Location / Form / notes / supplements) -----------
# The sheet carries, below the exercise block, a labelled "Location" row, a
# labelled "Form" row (self-assessed 1-10, Italian decimals like "6,5"), and a
# few UNLABELLED annotation rows mixing supplement shorthand (Cp/C/P) with free
# notes (Caldo, Stanco, Anavar…). Captured here as session metadata, typed where
# possible (see /CLAUDE.md → Data model).

def parse_form(v):
    """Self-assessed form 1-10; accepts int/float or 'x,y' Italian decimals."""
    if v is None: return None
    if isinstance(v, bool): return None
    if isinstance(v, (int, float)):
        f = float(v)
    else:
        s = str(v).strip().replace(",", ".")
        if not re.fullmatch(r"\d+(\.\d+)?", s): return None
        f = float(s)
    return f if 1 <= f <= 10 else None

# Supplement shorthand: the owner writes letters for what he took — c=creatine,
# p=protein, m=maca, a=aminos — as letter-sets (`Cp`, `Cpm`, `Cpa`),
# slash/space-separated (`C/P/M`, `Cr/pr`), or spelled out. `Co` is a typo for
# `Cp`. A cell can mix a supplement token with a free note (`Cp, Anavar`).
SUPP_LETTER = {"c": "creatine", "p": "protein", "m": "maca", "a": "aminos"}
SPELLED = {
    "cr": "creatine", "creat": "creatine", "creatina": "creatine", "creatine": "creatine",
    "pr": "protein", "prot": "protein", "prote": "protein", "protein": "protein",
    "proteina": "protein", "proteine": "protein",
    "maca": "maca",
    "amino": "aminos", "aminos": "aminos", "amini": "aminos", "aminoacidi": "aminos",
    "co": "creatine,protein",  # common typo for "Cp"
}

def token_supps(tok):
    """Supplements a single token denotes, or None if it isn't supplement shorthand."""
    low = tok.lower()
    if low in SPELLED: return set(SPELLED[low].split(","))
    if re.fullmatch(r"[cpma]+", low): return {SUPP_LETTER[ch] for ch in low}
    return None

def classify_annotations(cells):
    """Split annotation cells into typed supplements + a free-text note string."""
    supp, notes = set(), []
    for raw in cells:
        if raw is None: continue
        s = str(raw).strip()
        if not s: continue
        if re.fullmatch(r"\d+([.,]\d+)?", s):
            continue  # stray number that leaked into an annotation cell
        leftover = []
        for part in re.split(r"[\s/,]+", s):
            if not part: continue
            hit = token_supps(part)
            if hit:
                supp |= hit
            else:
                leftover.append(part)
        if leftover:
            notes.append(" ".join(leftover))
    order = {"creatine": 0, "protein": 1, "maca": 2, "aminos": 3}
    return sorted(supp, key=lambda x: order[x]), ", ".join(notes)

def parse_value(v, exercise):
    """Return dict(reps, weightKg, modifiers, struggle) or None."""
    if v is None: return None
    # Excel silently converted a typed integer into a date — the day is the value.
    if isinstance(v, datetime.datetime):
        return {"reps": v.day, "weightKg": None, "modifiers": [], "struggle": False}
    s = str(v).strip()
    if not s: return None
    # Bare 3-digit value = "RxW" shorthand typed without the x (e.g. 820 → 8×20).
    if re.fullmatch(r"\d{3}", s):
        val = int(s)
        return {"reps": val // 100, "weightKg": val % 100, "modifiers": [], "struggle": False}
    mods, struggle = [], False
    if any(e in s for e in ["\U0001F613", "\U0001F975", "\U0001F624"]): struggle = True
    low = s.lower()
    if "band" in low: mods.append("band_travel")
    if re.search(r"nk", s, re.I): mods.append("no_kip")
    if "trx" in low: mods.append("trx")
    if "full" in low: mods.append("full_rom")
    if exercise in PULL and re.search(r"\bl\b|\(l\)|\d+\s*l\b", s, re.I): mods.append("L_sit")
    if exercise in PULL and "trx" not in low and not rxw(s) and re.search(r"\d\s*[xX]", s):
        mods.append("wide_X")
    m = rxw(s)
    reps = weight = None
    if m:
        reps, weight = int(m[1]), int(m[2])          # reps x kg
    else:
        ap = re.match(r"\s*(\d+)\s*\+\s*(\d+)", s)
        if ap:
            reps = int(ap[1]) + int(ap[2])           # a+b = total reps
        else:
            d = re.search(r"(\d+)", s)
            if d: reps = int(d[1])
    if reps is None: return None
    # Implausible bodyweight rep count = a number embedded in a note (e.g. the
    # "180" in "Cavi180lb"); no real move exceeds ~120 bodyweight reps.
    if weight is None and reps > 120: return None
    return {"reps": reps, "weightKg": weight, "modifiers": mods, "struggle": struggle}

def main(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    today = datetime.date.today()
    sessions, templates = [], []
    ex_types = {}          # canonical name -> list of has_weight bools
    ex_seen_order = {}     # sheet -> ordered unique exercises

    for ws in wb.worksheets:
        if ws.max_row < 2:  # empty sheet
            continue
        workout = ws.title
        datecols = {c: parse_date(ws.cell(1, c).value)
                    for c in range(2, ws.max_column + 1)
                    if parse_date(ws.cell(1, c).value)}
        # session per date column
        col_sets = {c: [] for c in datecols}
        order = []
        round_counter = {}   # (col, exercise) -> round
        # Locate the labelled meta rows; annotation rows are everything below Form.
        loc_r = form_r = None
        for r in range(2, ws.max_row + 1):
            lab = ws.cell(r, 1).value
            lab = str(lab).strip() if lab is not None else ""
            if lab == "Location": loc_r = r
            elif lab == "Form": form_r = r
        # Annotation cells sit in the few unlabelled rows just below Form.
        anno_rows = range(form_r + 1, min(ws.max_row, form_r + 8) + 1) if form_r else range(0)
        for r in range(2, ws.max_row + 1):
            raw = ws.cell(r, 1).value
            if not raw or not str(raw).strip(): continue
            name = str(raw).strip()
            # Skip divider/label rows that carry no exercise name (e.g. 💦💦💦).
            if not re.search(r"[A-Za-z0-9]", name): continue
            # Skip non-exercise meta rows (form rating / self-assessment / gym).
            if name in ("Form", "Fitness level", "Location"): continue
            forced_base, forced_mod = FORCE_MOVE_MOD.get(name, (None, None))
            ex = canon(forced_base or name)
            if ex not in order: order.append(ex)
            for c, d in datecols.items():
                if d > today: continue          # clip stray future dates
                pv = parse_value(ws.cell(r, c).value, name)
                if not pv: continue
                mods = pv["modifiers"]
                if forced_mod and forced_mod not in mods:
                    mods = mods + [forced_mod]
                key = (c, ex)
                round_counter[key] = round_counter.get(key, 0) + 1
                ex_types.setdefault(ex, []).append(pv["weightKg"] is not None)
                col_sets[c].append({
                    "id": str(uuid.uuid4()), "exercise": ex,
                    "reps": pv["reps"], "weightKg": pv["weightKg"],
                    "round": round_counter[key], "modifiers": mods,
                    "struggle": pv["struggle"],
                })
        ex_seen_order[workout] = order
        for c, d in datecols.items():
            if d > today or not col_sets[c]: continue
            sess = {"id": str(uuid.uuid4()), "date": d.isoformat(),
                    "workout": workout, "sets": col_sets[c]}
            # Attach per-day metadata captured from the meta / annotation rows.
            loc = ws.cell(loc_r, c).value if loc_r else None
            loc = str(loc).strip() if loc is not None else ""
            if loc and loc != "Location":
                sess["location"] = loc
            form = parse_form(ws.cell(form_r, c).value) if form_r else None
            if form is not None:
                sess["form"] = form
            supp, note = classify_annotations(
                [ws.cell(r, c).value for r in anno_rows])
            if supp:
                sess["supplements"] = supp
            if note:
                sess["notes"] = note
            sessions.append(sess)

    # Reverse the canonicalization maps so each exercise carries its known aliases.
    alias_map = collections.defaultdict(list)
    for a, c in ALIASES.items():
        alias_map[c].append(a)
    for a, (base, _mod) in FORCE_MOVE_MOD.items():
        alias_map[base].append(a)
    exercises = [{"name": e, "canonicalName": e,
                  "type": "weighted" if sum(v) > len(v) / 2 else "bodyweight",
                  "aliases": sorted(alias_map.get(e, []))}
                 for e, v in sorted(ex_types.items())]
    templates = [{"name": wk, "exercises": [canon(x) for x in order]}
                 for wk, order in ex_seen_order.items()]

    out = {"exercises": exercises, "templates": templates, "sessions": sessions}
    json.dump(out, sys.stdout, ensure_ascii=False)
    print(f"\n# {len(sessions)} sessions, "
          f"{sum(len(s['sets']) for s in sessions)} sets, "
          f"{len(exercises)} exercises", file=sys.stderr)

if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "P90X.xlsx")
